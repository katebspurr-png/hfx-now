from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: .venv/bin/python -m pip install openpyxl"
    ) from exc

VALID_TIERS = {"high", "medium", "low"}
VALID_STATUS = {"active", "paused"}
VALID_POLL_RE = re.compile(r"^\s*(\d+)\s*(h|d)\s*$|^\s*weekly\s*$", re.IGNORECASE)


def _normalize_url(raw_value: str) -> str:
    raw = (raw_value or "").strip()
    if not raw:
        return ""
    if raw.startswith("@"):
        raw = raw[1:]
    if not raw.startswith(("http://", "https://")):
        raw = f"https://{raw}"
    parsed = urlparse(raw)
    host = parsed.netloc.lower().strip()
    path = parsed.path.strip()
    if not host:
        return ""
    return f"https://{host}{path}"


def _row_to_dict(headers: Iterable[object], values: Iterable[object]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for header, value in zip(headers, values):
        key = str(header).strip() if header is not None else ""
        if not key:
            continue
        out[key] = "" if value is None else str(value).strip()
    return out


def _extract_rows(xlsx_path: Path, sheet_name: Optional[str]) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    row_iter = ws.iter_rows(values_only=True)
    headers = next(row_iter, None)
    if not headers:
        return []
    rows: List[Dict[str, str]] = []
    for values in row_iter:
        row = _row_to_dict(headers, values)
        if any(val.strip() for val in row.values()):
            rows.append(row)
    return rows


def _write_report_csv(
    report_path: Path,
    errors: List[str],
    warnings: List[str],
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["severity", "message"])
        writer.writeheader()
        for issue in errors:
            writer.writerow({"severity": "error", "message": issue})
        for issue in warnings:
            writer.writerow({"severity": "warning", "message": issue})


def _validate_rows(rows: List[Dict[str, str]]) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    club_names = [row.get("Club Name", "").strip() for row in rows if row.get("Club Name", "").strip()]
    name_counts = Counter([name.lower() for name in club_names])
    duplicate_names = sorted([name for name, count in name_counts.items() if count > 1])
    for name in duplicate_names:
        warnings.append(f"Duplicate club name rows: {name}")

    for idx, row in enumerate(rows, start=2):
        club_name = row.get("Club Name", "").strip()
        website = row.get("Website", "").strip()
        instagram = row.get("Instagram", "").strip()
        facebook = row.get("Facebook", "").strip()
        tier = row.get("Activity Tier", "").strip().lower()
        status = row.get("Status", "").strip().lower()
        poll_frequency = row.get("Poll Frequency", "").strip()

        if not club_name:
            warnings.append(f"Row {idx}: Missing Club Name.")
            continue

        if not any([website, instagram, facebook]):
            errors.append(f"Row {idx} ({club_name}): No Website/Instagram/Facebook source provided.")

        if website:
            normalized = _normalize_url(website)
            if not normalized:
                errors.append(f"Row {idx} ({club_name}): Website value is malformed: {website}")

        if instagram and not (instagram.startswith("@") or "instagram.com" in instagram.lower()):
            warnings.append(f"Row {idx} ({club_name}): Instagram value unusual: {instagram}")
        if facebook and not (facebook.startswith("@") or "facebook.com" in facebook.lower()):
            warnings.append(f"Row {idx} ({club_name}): Facebook value unusual: {facebook}")

        if tier and tier not in VALID_TIERS:
            errors.append(
                f"Row {idx} ({club_name}): Invalid Activity Tier '{tier}'. Use high|medium|low."
            )
        if status and status not in VALID_STATUS:
            errors.append(
                f"Row {idx} ({club_name}): Invalid Status '{status}'. Use active|paused."
            )
        if poll_frequency and not VALID_POLL_RE.match(poll_frequency):
            warnings.append(
                f"Row {idx} ({club_name}): Poll Frequency '{poll_frequency}' not in suggested formats (6h, 24h, 7d, weekly)."
            )
    return errors, warnings


def _write_cleaned_csv(rows: List[Dict[str, str]], output_csv: Path) -> int:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fixes = 0
    cleaned: List[Dict[str, str]] = []

    for row in rows:
        out = dict(row)
        facebook = out.get("Facebook", "").strip()
        notes = out.get("Notes", "").strip()
        if facebook and "facebook.com" not in facebook.lower() and not facebook.startswith("@"):
            if "whatsapp.com" in facebook.lower():
                out["Facebook"] = ""
                out["Notes"] = (notes + " | WhatsApp group: " + facebook).strip(" |")
                fixes += 1
        cleaned.append(out)

    headers: List[str] = []
    for row in cleaned:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in cleaned:
            writer.writerow(row)
    return fixes


def validate_workbook(
    xlsx_path: Path,
    sheet_name: Optional[str],
    report_csv: Optional[Path],
    autofix_output_csv: Optional[Path],
) -> int:
    rows = _extract_rows(xlsx_path, sheet_name)
    errors, warnings = _validate_rows(rows)

    print(f"Workbook: {xlsx_path}")
    print(f"Rows checked: {len(rows)}")
    print("")
    print(f"Errors: {len(errors)}")
    for issue in errors[:100]:
        print(f"- {issue}")
    if len(errors) > 100:
        print(f"- ... and {len(errors) - 100} more errors")
    print("")
    print(f"Warnings: {len(warnings)}")
    for issue in warnings[:100]:
        print(f"- {issue}")
    if len(warnings) > 100:
        print(f"- ... and {len(warnings) - 100} more warnings")

    if report_csv:
        _write_report_csv(report_csv, errors, warnings)
        print("")
        print(f"Report CSV: {report_csv}")

    if autofix_output_csv:
        fixes = _write_cleaned_csv(rows, autofix_output_csv)
        print(f"Autofix output CSV: {autofix_output_csv}")
        print(f"Autofixes applied: {fixes}")

    return 1 if errors else 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate run-club workbook before import.")
    parser.add_argument("--xlsx", required=True, help="Path to workbook.")
    parser.add_argument("--sheet", default="", help="Sheet name (defaults to first).")
    parser.add_argument(
        "--report-csv",
        default="",
        help="Optional path to write validation findings as CSV.",
    )
    parser.add_argument(
        "--autofix-output-csv",
        default="",
        help="Optional path to write a cleaned CSV with safe autofixes.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")
    report_csv = Path(args.report_csv).expanduser().resolve() if args.report_csv else None
    autofix_output_csv = (
        Path(args.autofix_output_csv).expanduser().resolve() if args.autofix_output_csv else None
    )
    raise SystemExit(
        validate_workbook(
            xlsx_path=xlsx_path,
            sheet_name=args.sheet or None,
            report_csv=report_csv,
            autofix_output_csv=autofix_output_csv,
        )
    )


if __name__ == "__main__":
    main()
