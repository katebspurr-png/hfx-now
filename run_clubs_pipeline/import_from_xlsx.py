from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: .venv/bin/python -m pip install openpyxl"
    ) from exc

try:
    from run_clubs_pipeline.models import ClubSource
    from run_clubs_pipeline.source_registry import load_registry, write_registry
except ModuleNotFoundError:  # pragma: no cover - direct script execution fallback
    from models import ClubSource
    from source_registry import load_registry, write_registry


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "club"


def _normalize_url(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    raw = raw.replace(" ", "")
    if raw.startswith("@"):
        raw = raw[1:]
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    parsed = urlparse(raw)
    host = parsed.netloc.lower()
    path = parsed.path.rstrip("/")
    if not path and host in {"instagram.com", "www.instagram.com", "facebook.com", "www.facebook.com"}:
        return ""
    normalized = f"https://{host}{path}"
    if parsed.query:
        normalized += f"?{parsed.query}"
    return normalized


def _instagram_url(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    if raw.startswith("@"):
        return f"https://www.instagram.com/{raw[1:]}"
    url = _normalize_url(raw)
    if "instagram.com" in url:
        return url
    return ""


def _facebook_url(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    if raw.startswith("@"):
        return f"https://www.facebook.com/{raw[1:]}"
    url = _normalize_url(raw)
    if "facebook.com" in url:
        return url
    return ""


def _source_from_website(website_value: str) -> Tuple[str, str]:
    url = _normalize_url(website_value)
    if not url:
        return "", ""
    host = urlparse(url).netloc.lower()
    if "instagram.com" in host:
        return "instagram", _instagram_url(url)
    if "facebook.com" in host:
        return "facebook", _facebook_url(url)
    return "website", url


def _tier_from_row(row: Dict[str, str]) -> str:
    explicit = row.get("Activity Tier", "").strip().lower()
    if explicit in {"high", "medium", "low"}:
        return explicit
    notes = " ".join(
        [
            row.get("Type", ""),
            row.get("Pace", ""),
            row.get("Notes", ""),
        ]
    ).lower()
    if "coached" in notes or "performance" in notes:
        return "high"
    if "weekly" in notes or "multiple" in notes:
        return "high"
    if "check website" in notes or "seasonal" in notes:
        return "medium"
    return "medium"


def _status_from_row(row: Dict[str, str]) -> str:
    explicit = row.get("Status", "").strip().lower()
    return explicit if explicit in {"active", "paused"} else "active"


def _poll_frequency_from_row(row: Dict[str, str]) -> str:
    return row.get("Poll Frequency", "").strip()


def _build_sources_from_row(row: Dict[str, str]) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    website_kind, website_url = _source_from_website(row.get("Website", ""))
    if website_kind and website_url:
        out.append((website_kind, website_url))

    ig_url = _instagram_url(row.get("Instagram", ""))
    if ig_url:
        out.append(("instagram", ig_url))

    fb_url = _facebook_url(row.get("Facebook", ""))
    if fb_url:
        out.append(("facebook", fb_url))

    # De-dupe same source URL/type pairs.
    deduped: List[Tuple[str, str]] = []
    seen = set()
    for source_type, source_url in out:
        key = (source_type, source_url.lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append((source_type, source_url))
    return deduped


def _row_to_dict(headers: Iterable[object], values: Iterable[object]) -> Dict[str, str]:
    row: Dict[str, str] = {}
    for header, value in zip(headers, values):
        key = str(header).strip() if header is not None else ""
        if not key:
            continue
        row[key] = "" if value is None else str(value).strip()
    return row


def _extract_rows(xlsx_path: Path, sheet_name: Optional[str]) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    rows: List[Dict[str, str]] = []
    for values in rows_iter:
        row = _row_to_dict(headers, values)
        if not row.get("Club Name", "").strip():
            continue
        rows.append(row)
    return rows


def _index_existing(existing: List[ClubSource]) -> Dict[Tuple[str, str, str], ClubSource]:
    index: Dict[Tuple[str, str, str], ClubSource] = {}
    for src in existing:
        k = (src.club_name.lower(), src.source_type, src.source_url.lower())
        index[k] = src
    return index


def import_registry(xlsx_path: Path, registry_path: Path, sheet_name: Optional[str], replace_all: bool) -> Tuple[int, int]:
    rows = _extract_rows(xlsx_path, sheet_name)
    existing = [] if replace_all else load_registry(registry_path)
    existing_index = _index_existing(existing)

    output: List[ClubSource] = []
    if not replace_all:
        output.extend(existing)

    added = 0
    updated = 0
    used_ids = {src.club_id for src in output}

    for row in rows:
        club_name = row.get("Club Name", "").strip()
        if not club_name:
            continue
        tier = _tier_from_row(row)
        sources = _build_sources_from_row(row)
        for source_type, source_url in sources:
            key = (club_name.lower(), source_type, source_url.lower())
            maybe_existing = existing_index.get(key)
            if maybe_existing:
                maybe_existing.activity_tier = tier or maybe_existing.activity_tier
                maybe_existing.status = _status_from_row(row) or maybe_existing.status
                poll_frequency = _poll_frequency_from_row(row)
                maybe_existing.poll_frequency = poll_frequency or maybe_existing.poll_frequency
                note_bits = [maybe_existing.notes, row.get("Notes", "")]
                maybe_existing.notes = " | ".join([bit for bit in note_bits if bit]).strip(" |")
                updated += 1
                continue

            base_id = f"{_slugify(club_name)}-{source_type}"
            club_id = base_id
            i = 2
            while club_id in used_ids:
                club_id = f"{base_id}-{i}"
                i += 1
            used_ids.add(club_id)

            output.append(
                ClubSource(
                    club_id=club_id,
                    club_name=club_name,
                    source_type=source_type,
                    source_url=source_url,
                    activity_tier=tier,
                    poll_frequency=_poll_frequency_from_row(row),
                    status=_status_from_row(row),
                    last_checked_at="",
                    manual_review_due_at="",
                    notes=row.get("Notes", ""),
                )
            )
            added += 1

    # Stable order for predictable diffs and review.
    output_sorted = sorted(output, key=lambda s: (s.club_name.lower(), s.source_type, s.source_url.lower()))
    write_registry(registry_path, output_sorted)
    return added, updated


def main() -> None:
    parser = argparse.ArgumentParser(description="Import run club sources from xlsx into source_registry.csv")
    parser.add_argument("--xlsx", required=True, help="Path to Excel workbook")
    parser.add_argument("--sheet", default="", help="Sheet name (defaults to first sheet)")
    parser.add_argument(
        "--registry",
        default="data/run_clubs/source_registry.csv",
        help="Output registry CSV path",
    )
    parser.add_argument(
        "--replace-all",
        action="store_true",
        help="Replace registry rows with workbook-derived rows only",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    registry_path = Path(args.registry).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    added, updated = import_registry(
        xlsx_path=xlsx_path,
        registry_path=registry_path,
        sheet_name=args.sheet or None,
        replace_all=args.replace_all,
    )
    print(f"Imported workbook: {xlsx_path}")
    print(f"Registry written: {registry_path}")
    print(f"Rows added: {added}")
    print(f"Rows updated: {updated}")


if __name__ == "__main__":
    main()
